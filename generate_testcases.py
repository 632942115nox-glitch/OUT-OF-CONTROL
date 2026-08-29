"""Generate the structured test case Excel spreadsheet for Dungeon Guardian."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2D2D44", end_color="2D2D44", fill_type="solid")
p0_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
p1_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
p2_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
module_font = Font(name="Microsoft YaHei", size=11, bold=True, color="1A1A2E")
module_fill = PatternFill(start_color="C8C8E0", end_color="C8C8E0", fill_type="solid")
cell_font = Font(name="Microsoft YaHei", size=10)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)

# ── Sheets ──
# Sheet 1: 测试用例
ws = wb.active
ws.title = "测试用例"

headers = ["编号", "模块", "用例名称", "优先级", "前置条件", "测试步骤", "预期结果", "实际结果", "状态", "备注"]
col_widths = [10, 12, 28, 8, 30, 40, 40, 30, 10, 20]

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 28

# Test cases data
test_cases = [
    # ── Module 1: 状态机 ──
    ("TC-SM-001", "状态机", "正常启动流程 - 显示主菜单", "P0",
     "浏览器打开index.html",
     "1. 打开页面\n2. 观察主菜单",
     "显示主菜单（新游戏/继续/设置按钮），画布显示空白", "", "", ""),
    ("TC-SM-002", "状态机", "新游戏→游戏中 迁移", "P0",
     "在主菜单",
     "点击\"新游戏\"",
     "菜单消失，地图显示，HUD面板显示金币200、生命20", "", "", ""),
    ("TC-SM-003", "状态机", "游戏中→暂停→继续", "P0",
     "游戏中",
     "1. 按Space键\n2. 观察暂停弹窗\n3. 点击\"继续游戏\"",
     "暂停弹窗出现，游戏停止；点击继续后恢复", "", "", ""),
    ("TC-SM-004", "状态机", "暂停→返回菜单", "P1",
     "已暂停",
     "点击\"返回菜单\"",
     "回到主菜单，游戏状态重置", "", "", ""),
    ("TC-SM-005", "状态机", "生命归零→游戏结束", "P1",
     "游戏中，生命≤1",
     "让一个敌人到达终点",
     "显示游戏结束弹窗，包含统计数据（用时/波次/击杀等）", "", "", ""),
    ("TC-SM-006", "状态机", "通关全部波次→胜利", "P1",
     "通过第10波",
     "击杀最后一波所有敌人",
     "显示胜利弹窗，成就解锁提示", "", "", ""),
    ("TC-SM-007", "状态机", "快速连续点击\"重新开始\"", "P2",
     "游戏结束弹窗显示",
     "快速连续点击3次\"重新开始\"",
     "只启动一次新游戏，无异常弹窗或报错", "", "", ""),
    ("TC-SM-008", "状态机", "非法状态迁移（MENU直接到PAUSED）", "P2",
     "控制台",
     "执行 DG.GameState.transition('PAUSED')",
     "日志输出ERROR级别，状态不改变", "", "", ""),

    # ── Module 2: 地图 ──
    ("TC-MAP-001", "地图", "路径上不可放置塔", "P0",
     "新游戏",
     "选择弓箭塔，点击路径格子",
     "显示\"无法在此位置放置\"，塔未放置，金币未扣", "", "", ""),
    ("TC-MAP-002", "地图", "空地正常放置塔", "P0",
     "新游戏，金币≥50",
     "1. 点击弓箭塔\n2. 点击空地",
     "塔成功放置，金币扣除50", "", "", ""),
    ("TC-MAP-003", "地图", "同一位置不重复放置", "P1",
     "某位置已有塔",
     "选择另一种塔，点击该位置",
     "不会覆盖，而是选中已有塔显示信息面板", "", "", ""),
    ("TC-MAP-004", "地图", "地图边缘格子放置", "P1",
     "新游戏",
     "点击地图最边缘(0,0)、(19,14)等位置",
     "边界格子正常响应，可放置/选中", "", "", ""),
    ("TC-MAP-005", "地图", "点击地图外区域", "P1",
     "选择了塔类型",
     "点击画布外区域（侧边栏）",
     "塔未放置，hover高亮消失", "", "", ""),
    ("TC-MAP-006", "地图", "路径起终点标识", "P2",
     "新游戏",
     "观察地图左上和右下",
     "起点绿色\"入\"标识，终点红色\"核\"标识", "", "", ""),

    # ── Module 3: 塔系统 ──
    ("TC-TWR-001", "塔", "4种塔均可正常放置", "P0",
     "金币≥200",
     "每种塔各放置1个",
     "弓箭50/炮80/冰60/魔法70金币正确扣除", "", "", ""),
    ("TC-TWR-002", "塔", "金币不足时无法放置", "P0",
     "金币<50",
     "尝试放置弓箭塔",
     "提示\"金币不足\"，塔未放置", "", "", ""),
    ("TC-TWR-003", "塔", "塔升级 Lv.1→Lv.2→Lv.3", "P0",
     "放置弓箭塔，有足够金币",
     "1. 选中塔\n2. 点击升级\n3. 再次升级",
     "花费75+150，伤害15→25→40，范围120→140→160，攻速提升", "", "", ""),
    ("TC-TWR-004", "塔", "最高级无法升级", "P1",
     "Lv.3 塔",
     "选中塔，查看升级按钮",
     "按钮显示\"已达最高级\"，disabled状态", "", "", ""),
    ("TC-TWR-005", "塔", "出售返还60%金币（Lv.1）", "P1",
     "Lv.1弓箭塔（花费50）",
     "选中塔，点击出售",
     "返还30金币（50×0.6），塔消失，格子可重新放置", "", "", ""),
    ("TC-TWR-006", "塔", "升级后出售返还计算", "P1",
     "弓箭塔Lv.3（总花费275）",
     "出售该塔",
     "返还165金币（275×0.6）", "", "", ""),
    ("TC-TWR-007", "塔", "弓箭塔攻击蝙蝠（飞行单位）", "P1",
     "地图上有弓箭塔，蝙蝠出现",
     "观察弓箭塔是否攻击蝙蝠",
     "弓箭塔正常攻击蝙蝠", "", "", ""),
    ("TC-TWR-008", "塔", "炮塔不攻击飞行敌人", "P1",
     "地图上有炮塔，蝙蝠出现",
     "观察炮塔行为",
     "炮塔不攻击蝙蝠（飞行单位免疫地面AOE）", "", "", ""),
    ("TC-TWR-009", "塔", "冰塔减速效果", "P2",
     "冰塔攻击敌人",
     "观察被冰塔攻击的敌人",
     "敌人身上蓝色光晕，移动速度明显下降约60%", "", "", ""),
    ("TC-TWR-010", "塔", "魔法塔连锁闪电", "P2",
     "魔法塔附近有3+敌人",
     "观察魔法塔攻击",
     "闪电在敌人间跳跃，伤害逐跳递减25%", "", "", ""),

    # ── Module 4: 敌人 ──
    ("TC-ENM-001", "敌人", "敌人沿路径移动", "P0",
     "开始第1波",
     "不放置任何塔，观察敌人移动",
     "哥布林沿预设路径从起点走向终点", "", "", ""),
    ("TC-ENM-002", "敌人", "敌人到达终点扣生命", "P0",
     "不放置塔",
     "让敌人到达终点",
     "普通敌人扣1点，Boss扣5点", "", "", ""),
    ("TC-ENM-003", "敌人", "4种敌人属性差异", "P1",
     "完成前4波",
     "观察哥布林/兽人/巨魔/蝙蝠",
     "速度:蝙蝠>哥布林>兽人>巨魔；HP:巨魔>兽人>哥布林>蝙蝠", "", "", ""),
    ("TC-ENM-004", "敌人", "Boss特殊效果", "P1",
     "第5波开始",
     "观察Boss外观和行为",
     "体积更大、红色光环脉冲、HP条始终显示、扣5生命", "", "", ""),
    ("TC-ENM-005", "敌人", "巨魔生命回复", "P1",
     "巨魔受伤但未死",
     "停止攻击，观察数秒",
     "巨魔HP缓慢回升（每秒约2点）", "", "", ""),
    ("TC-ENM-006", "敌人", "蝙蝠免疫炮塔和冰塔", "P1",
     "有炮塔和冰塔，蝙蝠出现",
     "观察战斗日志",
     "蝙蝠不被炮塔和冰塔攻击，只被弓箭和魔法塔攻击", "", "", ""),
    ("TC-ENM-007", "敌人", "Tab切换后敌人恢复", "P2",
     "敌人正在移动",
     "切换Tab 5秒后切回",
     "敌人继续正常移动，无瞬移，帧率恢复", "", "", ""),
    ("TC-ENM-008", "敌人", "第10波大量敌人共存渲染", "P2",
     "第10波峰值",
     "观察约20+敌人同时存在",
     "所有敌人正常渲染和移动，无闪烁/重叠/消失", "", "", ""),

    # ── Module 5: 战斗 ──
    ("TC-CBT-001", "战斗", "单体伤害计算准确性", "P0",
     "弓箭塔Lv.1攻击哥布林(HP60)",
     "计算4次攻击的伤害",
     "每次15伤害，第4次击杀（15×4=60）", "", "", ""),
    ("TC-CBT-002", "战斗", "AOE范围伤害", "P0",
     "炮塔攻击，密集敌群",
     "观察爆炸范围内的敌人血条",
     "范围内所有敌人受到等量AOE伤害", "", "", ""),
    ("TC-CBT-003", "战斗", "冰塔减速叠加最小值", "P1",
     "多个冰塔攻击同一敌人",
     "观察减速效果",
     "减速叠加但最低不低于0.15倍基础速度", "", "", ""),
    ("TC-CBT-004", "战斗", "Boss减速抗性（50%）", "P1",
     "冰塔攻击Boss",
     "观察Boss减速效果",
     "Boss减速因子×0.5，持续时间×0.5", "", "", ""),
    ("TC-CBT-005", "战斗", "连锁闪电伤害递减", "P1",
     "魔法塔Lv.1攻击，3个敌人相邻",
     "观察连锁伤害数值",
     "第1目标20伤害→第2目标15→第3目标10（依次×0.75）", "", "", ""),
    ("TC-CBT-006", "战斗", "敌人死亡后不再受攻击", "P2",
     "多个塔集火同一敌人",
     "敌人HP归零瞬间观察后续弹丸",
     "弹丸命中死亡敌人不产生额外效果", "", "", ""),
    ("TC-CBT-007", "战斗", "冰塔对飞行敌人伤害减半", "P2",
     "冰塔攻击蝙蝠（HP30）",
     "观察伤害数值",
     "每次4点伤害（正常8点×50%）", "", "", ""),

    # ── Module 6: 经济 ──
    ("TC-ECO-001", "经济", "击杀敌人金币获取", "P0",
     "新游戏(200金币)，普通难度",
     "击杀一个哥布林",
     "金币增加10", "", "", ""),
    ("TC-ECO-002", "经济", "难度影响金币获取", "P0",
     "分别在简单/困难难度",
     "各击杀一个哥布林",
     "简单13金币(10×1.3)，困难8金币(10×0.8)", "", "", ""),
    ("TC-ECO-003", "经济", "波次完成奖励", "P1",
     "完成第1波",
     "检查金币变化",
     "额外获得30金币（20+10×1）波次奖励", "", "", ""),
    ("TC-ECO-004", "经济", "金币不能为负", "P1",
     "金币为0",
     "尝试放置任何塔",
     "提示\"金币不足\"，金币保持0", "", "", ""),
    ("TC-ECO-005", "经济", "大额金币不溢出", "P1",
     "控制台 DG.addGold(99999)",
     "观察金币显示",
     "显示正常，无科学计数法或溢出", "", "", ""),
    ("TC-ECO-006", "经济", "花费0或负数处理", "P2",
     "控制台 DG.Economy.spend(0)",
     "观察日志和金币",
     "日志\"无效花费: 0\"，金币不变", "", "", ""),

    # ── Module 7: 波次 ──
    ("TC-WAV-001", "波次", "波次逐波推进 1→10", "P0",
     "新游戏",
     "逐波点击\"开始下一波\"",
     "每波敌人数量/类型递增，计数器正确", "", "", ""),
    ("TC-WAV-002", "波次", "波次进行中不可开启下一波", "P0",
     "波次进行中",
     "尝试点击\"开始下一波\"",
     "按钮disabled，无法点击", "", "", ""),
    ("TC-WAV-003", "波次", "第5波Boss出现", "P1",
     "到达第5波",
     "开始第5波，观察生成顺序",
     "哥布林+兽人之后，Boss单独出现", "", "", ""),
    ("TC-WAV-004", "波次", "第10波双Boss混合", "P1",
     "到达第10波",
     "开始第10波",
     "2个Boss+巨魔+兽人+蝙蝠混合，共约20+敌人", "", "", ""),
    ("TC-WAV-005", "波次", "暂停期间波次停止", "P1",
     "波次进行中",
     "按Space暂停",
     "生成队列停止，所有敌人停止移动", "", "", ""),
    ("TC-WAV-006", "波次", "继续后波次恢复", "P1",
     "暂停中，波次未完成",
     "点击继续游戏",
     "敌人从暂停位置继续生成和移动", "", "", ""),
    ("TC-WAV-007", "波次", "通关后按钮状态", "P2",
     "全部10波完成",
     "观察\"开始下一波\"按钮",
     "显示\"全部波次完成\"，disabled", "", "", ""),
    ("TC-WAV-008", "波次", "快速双击防重复启动", "P2",
     "波次间",
     "快速双击\"开始下一波\"",
     "只启动一波，不重复", "", "", ""),

    # ── Module 8: 存档 ──
    ("TC-SAV-001", "存档", "保存后读档完整性", "P0",
     "放置3座塔，金币150，波次3",
     "1. 保存\n2. F5刷新\n3. 点击\"继续游戏\"",
     "塔位置/等级/类型一致，金币150，波次3", "", "", ""),
    ("TC-SAV-002", "存档", "无存档时\"继续\"隐藏", "P0",
     "清除localStorage",
     "打开游戏",
     "主菜单只有\"新游戏\"和\"设置\"，无\"继续游戏\"", "", "", ""),
    ("TC-SAV-003", "存档", "波次进行中保存恢复", "P0",
     "波次5进行中",
     "1. 保存\n2. 刷新\n3. 读档",
     "敌人位置和HP恢复，波次从中断点继续", "", "", ""),
    ("TC-SAV-004", "存档", "暂停状态存档恢复", "P1",
     "游戏中按Space暂停",
     "1. 暂停菜单存盘\n2. 刷新\n3. 读档",
     "恢复为暂停状态，可继续", "", "", ""),
    ("TC-SAV-005", "存档", "数据篡改防御（gold=\"abc\"）", "P1",
     "正常存档",
     "1. DevTools修改localStorage gold为\"abc\"\n2. 读档",
     "数据校验失败，显示错误日志，降级处理", "", "", ""),
    ("TC-SAV-006", "存档", "localStorage清除后读档", "P1",
     "正常存档",
     "1. 手动清除localStorage\n2. 点击\"继续\"",
     "显示\"无存档记录\"日志，开始新游戏", "", "", ""),
    ("TC-SAV-007", "存档", "通关后存档自动删除", "P1",
     "通关",
     "刷新后查看主菜单",
     "\"继续游戏\"按钮消失", "", "", ""),
    ("TC-SAV-008", "存档", "存档塔类型无效处理", "P2",
     "篡改存档塔type为\"invalid\"",
     "读档",
     "日志WARN提示无效类型，其他塔正常恢复", "", "", ""),

    # ── Module 9: 设置 ──
    ("TC-SET-001", "设置", "难度影响敌人HP", "P0",
     "分别设置简单/普通/困难",
     "各开一局，观察第1波哥布林HP",
     "简单42HP(60×0.7)，普通60，困难90(60×1.5)", "", "", ""),
    ("TC-SET-002", "设置", "音量滑块边界值", "P1",
     "设置面板",
     "拖动音量到0和100，尝试输入负数",
     "0%和100%可设置，超出范围被clamp", "", "", ""),
    ("TC-SET-003", "设置", "设置持久化（刷新不丢失）", "P1",
     "设置音量50，难度困难",
     "F5刷新后打开设置",
     "音量保持50%，难度保持困难", "", "", ""),
    ("TC-SET-004", "设置", "游戏中改难度不生效", "P1",
     "游戏中（难度普通）",
     "打开设置改为简单，继续游戏",
     "当前游戏难度不变（仅影响新游戏）", "", "", ""),
    ("TC-SET-005", "设置", "无效难度值处理", "P2",
     "控制台 DG.Settings.setDifficulty('impossible')",
     "观察日志",
     "日志WARN\"无效难度\"，难度不变", "", "", ""),

    # ── Module 10: 成就 ──
    ("TC-ACH-001", "成就", "首次击杀→第一滴血", "P0",
     "新游戏",
     "击杀第一个敌人",
     "屏幕右上角Toast\"成就解锁: 🩸 第一滴血\"", "", "", ""),
    ("TC-ACH-002", "成就", "金币1000→小有积蓄", "P1",
     "金币接近1000",
     "使金币达到1000",
     "弹出Toast", "", "", ""),
    ("TC-ACH-003", "成就", "通过第5波→半程冠军", "P1",
     "完成第5波",
     "击杀第5波所有敌人",
     "弹出Toast", "", "", ""),
    ("TC-ACH-004", "成就", "5座弓箭塔→神射手", "P1",
     "金币≥250",
     "放置第5座弓箭塔",
     "弹出Toast", "", "", ""),
    ("TC-ACH-005", "成就", "通关→地牢守卫者", "P1",
     "完成第10波",
     "击杀最终Boss",
     "弹出Toast+胜利弹窗", "", "", ""),
    ("TC-ACH-006", "成就", "成就不重复解锁", "P2",
     "已解锁\"第一滴血\"",
     "再次击杀敌人",
     "不弹出重复Toast", "", "", ""),

    # ── Module 11: UI ──
    ("TC-UI-001", "UI", "HUD面板数据实时更新", "P0",
     "游戏中",
     "击杀敌人，观察面板",
     "金币/击杀数立即更新", "", "", ""),
    ("TC-UI-002", "UI", "塔信息面板显示", "P0",
     "放置弓箭塔",
     "点击该塔",
     "侧边栏显示塔名/等级/伤害/范围/攻速/升级/出售", "", "", ""),
    ("TC-UI-003", "UI", "快捷键1-4选择塔", "P1",
     "游戏中",
     "分别按1/2/3/4键",
     "对应选择弓箭/炮/冰/魔法塔，鼠标跟随预览", "", "", ""),
    ("TC-UI-004", "UI", "Space暂停/继续", "P1",
     "游戏中",
     "按Space",
     "暂停弹窗出现；再按不会直接恢复（需按钮）", "", "", ""),
    ("TC-UI-005", "UI", "放置预览（绿/红）", "P1",
     "选择塔类型",
     "鼠标悬停空地 vs 路径",
     "空地绿色+范围圈，路径红色", "", "", ""),
    ("TC-UI-006", "UI", "右键取消选择", "P1",
     "正在放置塔或已选中塔",
     "右键点击画布",
     "取消所有选择状态，预览消失", "", "", ""),
    ("TC-UI-007", "UI", "Toast自动消失", "P2",
     "触发成就",
     "不做操作，等3秒",
     "Toast自动淡出消失", "", "", ""),
    ("TC-UI-008", "UI", "生命≤5红色警示", "P2",
     "生命降至5或以下",
     "观察HUD生命数值",
     "生命数值变为红色(danger class)", "", "", ""),

    # ── Module 12: 性能 ──
    ("TC-PERF-001", "性能", "空闲状态FPS≥58", "P0",
     "新游戏，无操作",
     "控制台 DG.getState().fps",
     "FPS≥58", "", "", ""),
    ("TC-PERF-002", "性能", "第10波满载FPS≥50", "P0",
     "第10波全敌人+10+塔攻击",
     "控制台 DG.getState().fps",
     "FPS≥50", "", "", ""),
    ("TC-PERF-003", "性能", "Tab切换不崩溃", "P1",
     "波次进行中",
     "切换Tab 30秒后切回",
     "游戏继续正常，无崩溃或画面异常", "", "", ""),
    ("TC-PERF-004", "性能", "30分钟内存稳定性", "P1",
     "新游戏",
     "运行30分钟，观察浏览器任务管理器内存",
     "内存增长<10MB，无持续泄漏趋势", "", "", ""),
]

# Write test cases
for i, tc in enumerate(test_cases, 2):
    priority = tc[3]
    for j, val in enumerate(tc, 1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.font = cell_font
        cell.alignment = wrap_align if j > 3 else center_align
        cell.border = thin_border

        # Priority background colors
        if priority == "P0":
            cell.fill = p0_fill
        elif priority == "P1":
            cell.fill = p1_fill
        elif priority == "P2":
            cell.fill = p2_fill

    ws.row_dimensions[i].height = 60

# Freeze header
ws.freeze_panes = "A2"
# Auto-filter
ws.auto_filter.ref = f"A1:J{len(test_cases)+1}"

# ── Sheet 2: 汇总统计 ──
ws2 = wb.create_sheet("汇总统计")

summary_data = [
    ("模块", "总用例", "P0", "P1", "P2", "通过率"),
    ("状态机", 8, 3, 3, 2, ""),
    ("地图", 6, 2, 3, 1, ""),
    ("塔", 10, 3, 5, 2, ""),
    ("敌人", 8, 2, 4, 2, ""),
    ("战斗", 7, 2, 3, 2, ""),
    ("经济", 6, 2, 3, 1, ""),
    ("波次", 8, 2, 4, 2, ""),
    ("存档", 8, 3, 4, 1, ""),
    ("设置", 5, 1, 3, 1, ""),
    ("成就", 6, 1, 4, 1, ""),
    ("UI", 8, 2, 4, 2, ""),
    ("性能", 4, 2, 2, 0, ""),
    ("合计", 84, 25, 42, 17, ""),
]

for i, row_data in enumerate(summary_data, 1):
    for j, val in enumerate(row_data, 1):
        cell = ws2.cell(row=i, column=j, value=val)
        cell.border = thin_border
        cell.alignment = center_align
        if i == 1:
            cell.font = header_font
            cell.fill = header_fill
        elif i == len(summary_data):
            cell.font = Font(name="Microsoft YaHei", size=11, bold=True)
            cell.fill = PatternFill(start_color="E8E8F0", end_color="E8E8F0", fill_type="solid")
        else:
            cell.font = cell_font

ws2.column_dimensions["A"].width = 12
for c in "BCDEF":
    ws2.column_dimensions[c].width = 10

# ── Sheet 3: Bug列表模板 ──
ws3 = wb.create_sheet("Bug记录")

bug_headers = ["Bug ID", "标题", "严重程度", "优先级", "模块", "前置条件", "复现步骤", "实际结果", "预期结果", "截图/日志", "状态", "发现日期"]
for i, h in enumerate(bug_headers, 1):
    cell = ws3.cell(row=1, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

ws3.column_dimensions["A"].width = 10
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 10
ws3.column_dimensions["D"].width = 8
ws3.column_dimensions["E"].width = 10
ws3.column_dimensions["F"].width = 25
ws3.column_dimensions["G"].width = 35
ws3.column_dimensions["H"].width = 25
ws3.column_dimensions["I"].width = 25
ws3.column_dimensions["J"].width = 20
ws3.column_dimensions["K"].width = 10
ws3.column_dimensions["L"].width = 12

ws3.freeze_panes = "A2"

# ── Save ──
output = r"C:\Users\Yan\WorkBuddy\2026-08-12-17-54-20\dungeon-guardian\DungeonGuardian_测试用例.xlsx"
wb.save(output)
print(f"Done! Saved to {output}")
print(f"Sheet 1 - 测试用例: {len(test_cases)} 条")
print("Sheet 2 - 汇总统计")
print("Sheet 3 - Bug记录模板")
